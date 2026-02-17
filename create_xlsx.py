#!/usr/bin/env python3
"""
Script to convert CSV entertainment ratings to XLSX format with formatting.
This creates a more visually appealing spreadsheet for easier reading.
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_xlsx_from_csv(csv_file, xlsx_file):
    """Convert CSV to XLSX with formatting."""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Entertainment Ratings"
    
    # Read CSV and write to XLSX
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Format header row
                if row_idx == 1:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Format favorite columns
                elif col_idx in [6, 8, 10, 12]:  # All Favorite columns
                    if value == "TRUE":
                        cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")
                
                # Format rating columns
                elif col_idx in [5, 7, 9, 11]:  # All Rating columns
                    cell.alignment = Alignment(horizontal="center")
                    try:
                        rating = float(value)
                        if rating >= 9.0:
                            cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                    except ValueError:
                        pass
                
                # Format year columns
                elif col_idx in [13, 14, 15]:  # Release_Year, End_Year, Next_Release columns
                    cell.alignment = Alignment(horizontal="center")
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
    
    # Adjust column widths
    column_widths = {
        'A': 12,  # Type
        'B': 25,  # Title
        'C': 18,  # Genre
        'D': 15,  # Where_To_Watch
        'E': 10,  # Sky_Rating
        'F': 12,  # Sky_Favorite
        'G': 10,  # Tay_Rating
        'H': 12,  # Tay_Favorite
        'I': 10,  # Landen_Rating
        'J': 12,  # Landen_Favorite
        'K': 10,  # Hanna_Rating
        'L': 12,  # Hanna_Favorite
        'M': 12,  # Release_Year
        'N': 10,  # End_Year
        'O': 14,  # Next_Release
        'P': 35,  # Notes
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the workbook
    wb.save(xlsx_file)
    print(f"Successfully created {xlsx_file}")

if __name__ == "__main__":
    create_xlsx_from_csv("entertainment_ratings.csv", "entertainment_ratings.xlsx")
